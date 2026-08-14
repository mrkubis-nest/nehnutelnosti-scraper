"""Export ponúk do Excelu (.xlsx) – čistá textová databáza bez obrázkov.

Súbor sa pri každom behu prepíše celý z databázy, takže vždy obsahuje
kompletný a aktuálny zoznam, nie len prírastok.
"""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path

from .export_local import cat_sk
from .models import Listing

log = logging.getLogger(__name__)

# (nadpis, šírka stĺpca)
COLUMNS = [
    ("Lokalita", 52),
    ("Inzerent", 28),
    ("Typ nehnuteľnosti", 30),
    ("Odkaz na ponuku", 18),
]


def write(path: str | Path, listings: list[Listing],
          new_ids: set[str] | None = None) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        log.error("Chýba balík openpyxl. Spusti: pip install openpyxl")
        return False

    new_ids = new_ids or set()
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Ponuky"

    header_fill = PatternFill("solid", fgColor="12263F")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    new_fill = PatternFill("solid", fgColor="E8F5EE")
    link_font = Font(color="0563C1", underline="single")

    for col, (title, width) in enumerate(COLUMNS, start=1):
        c = ws.cell(row=1, column=col, value=title)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 26

    for i, l in enumerate(listings, start=2):
        is_new = l.id in new_ids

        values = [
            l.location,
            l.advertiser_name,
            cat_sk(l),
            "otvoriť ponuku",
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row=i, column=col, value=value)

        link = ws.cell(row=i, column=4)
        link.hyperlink = l.url
        link.font = link_font

        if is_new:
            for col in range(1, len(COLUMNS) + 1):
                ws.cell(row=i, column=col).fill = new_fill

    last = len(listings) + 1
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{max(last, 2)}"

    _summary_sheet(wb, listings)

    try:
        wb.save(p)
    except PermissionError:
        log.error("Nedá sa zapísať %s – máš ten súbor otvorený v Exceli? "
                  "Zavri ho a spusti znova.", p)
        return False
    except OSError as exc:
        log.error("Zápis Excelu zlyhal (%s): %s", p, exc)
        return False

    log.info("Excel: %s (%d ponúk, %d nových)", p.resolve(), len(listings), len(new_ids))
    return True


def _summary_sheet(wb, listings: list[Listing]) -> None:
    """Druhý hárok so súhrnom – koľko ponúk kde a akého typu."""
    from openpyxl.styles import Font
    from collections import Counter

    ws = wb.create_sheet("Prehľad")
    bold = Font(bold=True)
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 14

    row = 1
    ws.cell(row=row, column=1, value="Súhrn").font = Font(bold=True, size=14)
    row += 2
    for label, value in [
        ("Ponúk celkom", len(listings)),
        ("Aktualizované", datetime.now().strftime("%d.%m.%Y %H:%M")),
    ]:
        ws.cell(row=row, column=1, value=label).font = bold
        ws.cell(row=row, column=2, value=value)
        row += 1

    for title, counter in [
        ("Podľa okresu", Counter(l.district or "–" for l in listings)),
        ("Podľa kategórie", Counter(cat_sk(l) for l in listings)),
    ]:
        row += 1
        ws.cell(row=row, column=1, value=title).font = Font(bold=True, size=12)
        row += 1
        ws.cell(row=row, column=1, value="Položka").font = bold
        ws.cell(row=row, column=2, value="Počet").font = bold
        row += 1
        for key, count in counter.most_common():
            ws.cell(row=row, column=1, value=key)
            ws.cell(row=row, column=2, value=count)
            row += 1
